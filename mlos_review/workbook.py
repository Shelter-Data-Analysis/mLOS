"""Every table this package builds, written out as a spreadsheet.

A slide shows the rows that fit. This is where the rest of them go, so that a
table the deck trimmed to five levels is still available in full, and so that a
number a reader wants to check is one they can select rather than retype.

**Not the R workbook.** `results/analysis_results.xlsx` is the analysis,
rendered from `results.json` by mlos_excel_export.R, and nothing here touches
it. This file holds what THIS package assembled: the same measures arranged
the way the deck arranges them, plus the comparisons the deck derives, which
the analysis workbook has no section for by design.

**What gets a sheet.** Every table a build actually creates, with three
exclusions:

  - a table that never reaches a reader, existing only as a step between two
    frames;
  - a table that is merely a VIEW of another table already written here: a row
    selection, a column subset, a transposition. `highlights_table` picks four
    rows out of `full_table`, and writing both would put the same numbers in
    two sheets while implying the shorter one had a source of its own;
  - a table that is another one under a different name, such as the
    competing-risk teaser, which IS the whole-sample stratum's table.

A table that COMBINES information from different places gets a sheet even when
its parts appear elsewhere. The three Cox sheets are the case: each fit is its
own table, and the comparison between them is a third, which is not a view of
either. The two ratio panels are the same case one step wider, and they are
where the Weibull fits and the Kaplan-Meier ratio are written out.

**One sheet per table, unless the columns are identical.** The default is a
sheet each. Where a family of tables is the same table computed once per
stratifier, so that the column headings match exactly, they are stacked on one
sheet, each under its own heading row, separated by a blank line. That is the
grouping used, because it is where a reader scanning down a sheet meets the
same columns throughout. Anything less certain gets its own sheet, which is
safe to split further and hard to get wrong.

**Numbers stay numbers.** A cell holds the value at full precision and carries
an Excel number format that says how to show it; a percentage is stored as the
fraction the bundle holds and displayed with a percent format. So a reader can
change the decimals, sort a column, or chart it, none of which a pre-rounded
string allows. Headers and rounding come from the same two helpers the pptx
renderer uses, in blocks.py, so a column cannot be headed one way on a slide
and another way here.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from mlos_review.blocks import (
    Format,
    Table,
    aj_levels_table,
    cell_format,
    cox_fit_table,
    cox_full_comparison_table,
    full_table,
    header_text,
    level_counts_table,
    ratio_panel_full_table,
    requires_aj,
    requires_full_table,
    requires_level_counts,
    requires_workload,
    workload_columns,
    workload_full_table,
    requires_sample,
    requires_study_window,
    sample_table,
    study_window_table,
)
from mlos_review.bundle import Bundle
from mlos_review.names import Vocabulary, capitalize_first
from mlos_review.regression import (comparison as cox_comparison, pooled,
                                    stratified, hazard_ratio_panel,
                                    los_ratio_panel, HR_SERIES, LOS_SERIES)

# Excel's own limits and conventions, not ours: 31 characters, and none of
# these characters, or the file will not open.
SHEET_NAME_MAX = 31
SHEET_NAME_BANNED = set(r"[]:*?/\\")

# What an empty cell says. A slide writes "n/a" so a reader is not left
# wondering whether a blank is a missing number or a missing row; a spreadsheet
# writes nothing, because a text cell in a numeric column is the thing that
# breaks a sort, and the emptiness carries the same meaning.
EMPTY_CELL = None


@dataclass
class Sheet:
    """One worksheet: a name, and the tables stacked on it in order."""

    name: str
    tables: list[Table] = field(default_factory=list)


def _stratifier_family(builder, bundle: Bundle, predicate) -> list[Table]:
    """One table per stratifier that qualifies, in the bundle's own order.

    The families this collects are exactly the ones the stacking rule is for:
    the same table computed per stratifier, so the columns match by
    construction rather than by coincidence.
    """
    return [builder(bundle, stratifier)
            for stratifier in bundle.stratifiers() if predicate(bundle, stratifier)]


def sheets(bundle: Bundle, vocab: Vocabulary) -> list[Sheet]:
    """Which tables this run produces, and how they are grouped onto sheets.

    The content decision, kept apart from `write`, which knows only how to put
    a frame into cells. A run missing a section contributes no sheet for it
    rather than an empty one, the same way a rule declines a slide.
    """
    out: list[Sheet] = []

    if requires_study_window(bundle):
        out.append(Sheet("Study_Window", [study_window_table(bundle)]))
    if requires_sample(bundle):
        out.append(Sheet("Sample", [sample_table(bundle, vocab)]))

    stays = _stratifier_family(lambda b, s: level_counts_table(b, s, vocab),
                               bundle, requires_level_counts)
    if stays:
        out.append(Sheet("Stays_By_Stratum", stays))

    los = _stratifier_family(full_table, bundle, requires_full_table)
    if los:
        out.append(Sheet("LOS_By_Stratum", los))

    # Every workload column at once, including the two the slides leave out:
    # the study-window animal-day total, which on a slide put a second scale of
    # the census beside it, and whichever of the three sections a stratifier
    # could not fill. A column not worth a third of a slide is still worth a
    # cell to a reader checking the arithmetic, and the width that made it a
    # poor slide column costs nothing here.
    workload_strata = [s for s in bundle.stratifiers()
                       if requires_workload(bundle, s, "census")]
    if workload_strata:
        # One agreed column set across the family, so the tables stack: the
        # stratifiers differ only in whether shares mean anything for them, and
        # splitting the sheet over that would separate tables a reader wants to
        # scan against each other.
        agreed = workload_columns(bundle, workload_strata)
        out.append(Sheet("Workload_By_Stratum",
                         [workload_full_table(bundle, s, vocab, agreed)
                          for s in workload_strata]))

    outcomes = _stratifier_family(lambda b, s: aj_levels_table(b, s, vocab),
                                  bundle, requires_aj)
    if outcomes:
        out.append(Sheet("Outcomes_By_Stratum", outcomes))

    # The two fits get a sheet each and their comparison a third. The first two
    # are not views of the third: each is a regression in its own right, and a
    # reader checking one hazard ratio should not have to read it out of a
    # table about something else.
    pooled_frame, stratified_frame = pooled(bundle), stratified(bundle)
    if not pooled_frame.empty:
        out.append(Sheet("Cox_Pooled", [
            cox_fit_table(pooled_frame, "pooled Cox regression")]))
    if not stratified_frame.empty:
        out.append(Sheet("Cox_Stratified", [
            cox_fit_table(stratified_frame, "stratified Cox regressions")]))
        out.append(Sheet("Cox_Comparison", [
            cox_full_comparison_table(cox_comparison(pooled_frame, stratified_frame))]))

    # The two ratio panels, which are where the Weibull fits and the
    # Kaplan-Meier ratio reach a reader as numbers at all: nothing above holds
    # either, and the slides that do show three estimates with their intervals
    # left to the whiskers. Each is one sheet rather than one per stratifier,
    # matching the Cox sheets, whose (stratifier, level) index these share.
    #
    # A panel holding nothing anywhere is a run that fitted none of its three
    # readings, and is skipped: an empty sheet reads as a missing number rather
    # than as an analysis that was never run. A panel holding some of them is
    # written with the rest blank, which is what a reader needs to see to know
    # which of the three the run has.
    for name, title, panel, series in (
            ("HR_Panel", "hazard ratios, three readings",
             hazard_ratio_panel(bundle), HR_SERIES),
            ("LOS_Ratio_Panel", "length-of-stay ratios, three readings",
             los_ratio_panel(bundle), LOS_SERIES)):
        if panel.empty or not panel.notna().any().any():
            continue
        out.append(Sheet(name, [ratio_panel_full_table(panel, series, title)]))

    return out


def _excel_format(spec: Format) -> str:
    """A Format as an Excel number-format string.

    The percent format is what makes `percent` a display choice rather than a
    stored one: the cell keeps the fraction, and Excel multiplies it for show,
    so changing the format never changes the data.
    """
    decimals = f"0.{'0' * spec.decimals}" if spec.decimals else "0"
    return f"{decimals}%" if spec.percent else decimals


def _rounds_to_zero(value: float, spec: Format) -> bool:
    """Whether this format would display a nonzero value as zero.

    The case is a p-value of 3.7e-06 under four decimals, which shows as
    "0.0000" and reads as an exact zero rather than as a number too small for
    the column. The cell still holds the value, but a sheet is read as printed,
    and a Cox p-value displayed as zero is the one number here a reader might
    quote. Such a cell falls back to General, where Excel shows 3.67183E-06.
    """
    if value == 0:
        return False
    scaled = abs(value) * (100 if spec.percent else 1)
    return scaled < 0.5 * 10 ** -spec.decimals


def _index_headers(frame: pd.DataFrame) -> list[str]:
    """Headers for the index columns, from the index's own names.

    A frame indexed by level yields "Level"; the regression frames, indexed by
    stratifier and level, yield both. An unnamed index yields a blank, which is
    what the stratum tables want: their first column holds level names under a
    heading the sheet title has already given.
    """
    names = list(frame.index.names)
    return [capitalize_first(name) if name else "" for name in names]


def _write_table(worksheet, table: Table, vocab: Vocabulary, row: int,
                 heading: bool) -> int:
    """Put one table on a sheet starting at `row`. Returns the next free row.

    The heading row appears only when a sheet carries more than one table,
    where it is what says which stratum a block of rows belongs to.
    On a sheet of one, the sheet's own name says it and a heading would repeat
    it directly above itself.
    """
    frame = table.df
    if heading:
        worksheet.cell(row=row, column=1, value=capitalize_first(table.title)).font = (
            worksheet.cell(row=row, column=1).font.copy(bold=True))
        row += 1

    index_headers = _index_headers(frame)
    for offset, text in enumerate(index_headers):
        worksheet.cell(row=row, column=1 + offset, value=text).font = (
            worksheet.cell(row=row, column=1 + offset).font.copy(bold=True))
    for offset, measure in enumerate(frame.columns):
        cell = worksheet.cell(row=row, column=1 + len(index_headers) + offset,
                              value=header_text(table, vocab, measure, short=False))
        cell.font = cell.font.copy(bold=True)
    row += 1

    for key, record in frame.iterrows():
        parts = key if isinstance(key, tuple) else (key,)
        for offset, part in enumerate(parts):
            worksheet.cell(row=row, column=1 + offset, value=str(part))
        for offset, measure in enumerate(frame.columns):
            column = 1 + len(index_headers) + offset
            value = record[measure]
            cell = worksheet.cell(row=row, column=column)
            if pd.isna(value):
                cell.value = EMPTY_CELL
                continue
            if isinstance(value, str):
                cell.value = value
                continue
            # The stored value, unrounded; the format decides the display.
            cell.value = float(value)
            spec = cell_format(table, key, measure)
            if not _rounds_to_zero(cell.value, spec):
                cell.number_format = _excel_format(spec)
        row += 1

    for note in table.footnotes:
        worksheet.cell(row=row, column=1, value=note).font = (
            worksheet.cell(row=row, column=1).font.copy(italic=True))
        row += 1

    return row


def _column_widths(worksheet) -> None:
    """Widen each column to its longest cell, within reason.

    A default-width column truncates "Per resident past days restricted
    median" to something unreadable, and a reader who has to widen columns
    before they can read the sheet will conclude the sheet was not meant to be
    read.
    """
    from openpyxl.utils import get_column_letter

    for column in worksheet.columns:
        longest = max((len(str(cell.value)) for cell in column if cell.value is not None),
                      default=0)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = (
            min(max(longest + 2, 10), 42))


def write(sheet_list: list[Sheet], path: str | Path, vocab: Vocabulary) -> Path | None:
    """Write the sheets to `path`. Returns where, or None when there are none.

    A workbook with no sheets is not a valid file, so a run that produced no
    tables at all writes nothing rather than something that will not open.
    """
    if not sheet_list:
        return None

    from openpyxl import Workbook

    book = Workbook()
    book.remove(book.active)
    for sheet in sheet_list:
        worksheet = book.create_sheet(sheet_name(sheet.name))
        row = 1
        for index, table in enumerate(sheet.tables):
            if index:
                row += 1   # the one-line gap between stacked tables
            row = _write_table(worksheet, table, vocab, row,
                               heading=len(sheet.tables) > 1)
        _column_widths(worksheet)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def sheet_name(name: str) -> str:
    """A name Excel will accept: no banned characters, no more than 31 of them."""
    cleaned = "".join("_" if character in SHEET_NAME_BANNED else character
                      for character in name)
    return cleaned[:SHEET_NAME_MAX]
